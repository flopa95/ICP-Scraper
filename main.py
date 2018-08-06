from bs4 import BeautifulSoup
import requests
import re
import openpyxl

document_location = "C:/Users/Florian Parzhuber/Desktop/ICP Check/KaylaTop20List.xlsx"
icp_test_pages = ["krustudios.com","legoland.com.my","simedarby.com","flystudio.my","udrive-media.com","ytl.com","parkson.com.my","primeworks.com.my","streamline-studios.com","aeu.edu.my","frogasia.com","thestar.com.my","orangesoft.com.my","ximnet.com.my","oum.edu.my","nry.com.my","tunehotels.com","malindoair.com","pttplc.com","2spotstudio.com","areeya.co.th","autoalliance.co.th","bangkokpost.com","kingpower.com","themallgroup.com","scg.com","workpoint.co.th","siamsport.co.th","nokair.com","vietjetair.com","lionairthai.com","nokscoot.com","thaismileair.com","thecabinchiangmai.com","theerawan.com","thaitrade.com","muaythairadio.com","dotpropertygroup.com","themonkstudio.com","mono.co.th","bumrungrad.com","dwp.com"]

counter = len(icp_test_pages)

wb = openpyxl.Workbook()
ws = wb.active

for i in range(counter):

    ## ADD A SECTION TO MODIFY THE DOMAIN FOR 6 SCENARIOS ##

    test_page1 = "http://" + icp_test_pages[i] 
    test_page2 = "https://" + icp_test_pages[i]
    test_page3 = "http://" + icp_test_pages[i] + ".cn"
    test_page4 = "https://" + icp_test_pages[i] + ".cn"
    test_page5 = "http://" + icp_test_pages[i].rsplit(".",1)[0] + ".cn"
    test_page6 = "https://" + icp_test_pages[i].rsplit(".", 1)[0] + ".cn"

    test_page_list = [test_page1,test_page2,test_page3,test_page4,test_page5,test_page6]

    ########################################################

    ws.cell(row = 1, column = 1).value ="Domain" 
    ws.cell(row = 1, column = 2).value ="ICP (Y/N)" 
    #ws.cell(row = 1, column = 3).value ="Found on these domains:"
    

    #ws.cell(row = 1, column = 2).value ="Domain" 
    #len(ws['C'])

    #ws.cell(row = len(ws['A']), column = 1).value = icp_test_pages[i]

    for x in range (6):        
        test_page = test_page_list[x]
        print(test_page)

        try:
            page = requests.get(test_page)
            soup = BeautifulSoup(page.content,"html.parser")

            if (bool(soup.find_all(text=re.compile("ICP"))) == True):
                ws.cell(row = len(ws['A'])+1, column = 1).value = test_page
                ws.cell(row = len(ws['A']), column = 2).value = "Y"
                print ("Most likely has an ICP")
                wb.save(document_location)
                #print (soup.find_all(text=re.compile("京ICP证")))    
            else:
                ws.cell(row = len(ws['A'])+1, column = 1).value = test_page
                ws.cell(row = len(ws['A']), column = 2).value = "N"
                print("Most likely has no ICP")
                wb.save(document_location)
    
        except:
            ws.cell(row = len(ws['A'])+1, column = 1).value = test_page
            ws.cell(row = len(ws['A']), column = 2).value = "Invalid domain"
            print("invalid domain, continue")
            wb.save(document_location)


wb.save(document_location)



